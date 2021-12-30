from Converter import Converter
from Checker import Checker
import random
import math
import openpyxl
import matplotlib.pyplot as plt


"""
Nome: Miguel Filipe Rodrigues Almeida de Matos Fazenda
Nº 110877
Curso: Mestrado em Engenharia Informática

Fixed Probability Counter: 1/16
Decreasing probability Counter: 1/sqrt(2)^k (1 sobre raíz quadrada de 2, com denominador elevado a k)
"""


# funções que contam os caracteres
def exact_count(counter, name):
    with open(name, "r", encoding="utf-8") as file:
        for line in file:
            for char in line:
                if char.isalpha():
                    if char in counter:  # se está no contador
                        counter[char] += 1

                    else:  # se não está no contador
                        counter[char] = 1

def fixed_count(counter, name):
    chance = 1/16

    with open(name, "r", encoding="utf-8") as file:
        for line in file:
            for char in line:
                if char.isalpha():
                    if char in counter:
                        if random.random() <= chance:  # random entre [0.0, 1.0[
                            counter[char] += 1

                    else:  # se não está no contador
                        if random.random() <= chance:  # conta adição
                            counter[char] = 1

                        else:  # não conta adição
                            counter[char] = 0 

def decreasing_count(counter, name):
    chance = 1/math.pow(math.sqrt(2), 0)  # cont = 0 -> k = 0 // cont = 1 -> k = 1 // cont = 2 -> k = 2 // etc

    with open(name, "r", encoding="utf-8") as file:
        for line in file:
            for char in line:
                if char.isalpha():
                    if char in counter:
                        if random.random() <= chance:  # random entre [0.0, 1.0[
                            counter[char] += 1
                            chance = 1/math.pow(math.sqrt(2), counter[char])

                    else:  # se não está no contador
                        if random.random() <= chance:  # conta adição
                            counter[char] = 1
                            chance = 1/math.pow(math.sqrt(2), counter[char])

                        else:  # não conta adição
                            counter[char] = 0


# funções de escrita para o ficheiro xls
# calcula também os erros absolutos e relativos
def write_info(exact, fixed, decreasing, name):
    try:
        wb = openpyxl.load_workbook("resultados/results.xlsx")  # carrega workbook(objeto xls)
        sheet = wb.create_sheet(name)  # cria nova sheet, fica no último índice
    except:
        wb = openpyxl.Workbook()  # cria workbook(objeto xls)
        sheet = wb.active  # sheet no índice 0
        sheet.title = name
    
    # ordena pelo primeiro valor do tuplo
    exact_list = sorted(exact.items(), key=lambda x: x[0])  
    fixed_list = sorted(fixed.items(), key=lambda x: x[0])
    dec_list = sorted(decreasing.items(), key=lambda x: x[0])

    sheet["A1"] = "Letters"
    sheet["B1"] = "Exact"

    sheet["D1"] = "Fixed (1/16)"
    sheet["E1"] = "Abs. Err."
    sheet["F1"] = "Rel. Err."

    sheet["H1"] = "Decreasing (1/(√2)^k)"
    sheet["I1"] = "Abs. Err."
    sheet["J1"] = "Rel. Err."

    for y in range(len(exact_list)):
        sheet["A" + str(y + 2)] = exact_list[y][0]  # letras
        sheet["B" + str(y + 2)] = exact_list[y][1]  # exact

        sheet["D" + str(y + 2)] = fixed_list[y][1]  # fixed
        sheet["E" + str(y + 2)] = exact_list[y][1] - fixed_list[y][1]  # absoluto
        sheet["F" + str(y + 2)] = (exact_list[y][1] - fixed_list[y][1]) / exact_list[y][1]  # relativo

        sheet["H" + str(y + 2)] = dec_list[y][1]    # decreasing
        sheet["I" + str(y + 2)] = exact_list[y][1] - dec_list[y][1]  # absoluto
        sheet["J" + str(y + 2)] = (exact_list[y][1] - dec_list[y][1]) / exact_list[y][1]  # relativo

    wb.save("resultados/results.xlsx")


# função que mostra e guarda os gráficos
def bar_plot(exact, fixed, decreasing, name):

    # legendas do grafo
    plt.figure(figsize = (16, 9))
    plt.xlabel("Letters")
    plt.ylabel("Counters")
    plt.title("Comparisons between exact, fixed K and decreasing K counters")

    # processa informação do exact counter
    exact_items = sorted(exact.items(), key=lambda x: x[0])
    exact_chars = []
    exact_counts = []

    for x in exact_items:
        exact_chars.append(x[0])
        exact_counts.append(math.log(x[1]))

    # processa informação do fixed
    fixed_items = sorted(fixed.items(), key=lambda x: x[0])
    fixed_chars = []
    fixed_counts = []

    for x in fixed_items:
        fixed_chars.append(x[0])
        try:
            fixed_counts.append(math.log(x[1]))
        except:  # log(0) é impossível
            fixed_counts.append(0)

    # processa informação do decereasing counter
    dec_items = sorted(decreasing.items(), key=lambda x: x[0])
    dec_chars = []
    dec_counts = []

    for x in dec_items:
        dec_chars.append(x[0])
        try:
            dec_counts.append(math.log(x[1]))
        except:
            dec_counts.append(0)

    # cria barras do grafo
    plt.bar(exact_chars, exact_counts, color="r", width=0.2, edgecolor="grey", label="Exact counter")
    plt.bar(fixed_chars, fixed_counts, color="g", width=0.2, edgecolor="grey", label="Approx. counter w/ 1/16 probability")
    plt.bar(dec_chars, dec_counts, color="b", width=0.2, edgecolor="grey", label="Approx. counter w/ decreasing probability")
    plt.xticks()

    plt.legend()
    plt.savefig("resultados/" + name + ".png")
    # plt.show()

    pass

def compare_letter(letter, top5):
    flag = False
    for x in top5:
        if letter != x[0]:
            flag = True

# função que conta se as 5 letras mais utilizadas são as mesmas e se são pela mesma ordem
def verify_letters(checker, exact_counter, fixed_counter, dec_counter):

    # ordena listas por quantidade de letras
    exact_list = sorted(exact_counter.items(), key=lambda x: x[1], reverse=True)
    fixed_list = sorted(fixed_counter.items(), key=lambda x: x[1], reverse=True)
    dec_list = sorted(dec_counter.items(), key=lambda x: x[1], reverse=True)
        
    # 5 letras mais utilizadas por ordem
    exact_top5 = []  
    fixed_top5 = []
    dec_top5 = []
    for x in range(5):
        exact_top5.append(exact_list[x][0])
        fixed_top5.append(fixed_list[x][0])
        dec_top5.append(dec_list[x][0])

    # compara fixed e dec top5 com exact_top5 em termos de letras
    fixed_letras_iguais = 0
    dec_letras_iguais = 0
    for x in range(5):
        if fixed_top5[x] in exact_top5:
            fixed_letras_iguais += 1
        if dec_top5[x] in exact_top5:
            dec_letras_iguais += 1
    
    if fixed_letras_iguais != 5:
        checker.fixed_same_letters += 1
    if dec_letras_iguais != 5:
        checker.dec_same_letters += 1

    # compara fixed e dec top5 com exact_top5 em termos de ordem
    fixed_flag = False
    dec_flag = False
    for x in range(5):
        if fixed_top5[x] != exact_top5[x]:
            fixed_flag = True
        if dec_top5[x] != exact_top5:
            dec_flag = True

    if fixed_flag:
        checker.fixed_order += 1
    if dec_flag:
        checker.dec_order += 1



if __name__ == "__main__":

    # só no caso de os ficheiros ainda n terem sido convertidos
    # comentar linha depois de convertidos para poupar o processador
    # Converter().small2big()

    random.seed()  # inicia random generator

    # checkers para verificar letras e ordem das letras
    en_checker = Checker()
    fr_checker = Checker()

    """versão inglesa"""
    exact_counter = {}
    fixed_counter = {}
    decreasing_counter = {}
    quant = 2000  # quantidde de iterções para verificar letras e ordem das letras

    # conta letras
    exact_count(exact_counter, "livros/3_MUSK_EN.txt")  # exact counter
    for x in range(quant):
        # limpra contadores
        fixed_counter.clear()
        decreasing_counter.clear()

        fixed_count(fixed_counter, "livros/3_MUSK_EN.txt")  # fixed counter with k value
        decreasing_count(decreasing_counter, "livros/3_MUSK_EN.txt")  # decreasing counter with increasing k value
        
        # verifica letras e ordem das letras são as mesmas da contagem exata
        verify_letters(en_checker, exact_counter, fixed_counter, decreasing_counter)
    
    # mostra chances
    print("---------------------Versão inglesa---------------------")
    print("Para approximate counting with fixed chance:")
    print(f"    chance de as letras serem diferentes: {(en_checker.fixed_order / quant) * 100}%")
    print(f"    chance de ordem das letras ser diferente: {(en_checker.fixed_same_letters / quant) * 100}%")
    print("Para approximate counting with decreasing chance:")
    print(f"    chance de as letras serem diferentes: {(en_checker.dec_order / quant) * 100}%")
    print(f"    chance de ordem das letras ser diferente: {(en_checker.dec_same_letters / quant) * 100}%")
    
    # escreve informação no xlsx
    write_info(exact_counter, fixed_counter, decreasing_counter, "en_version")

    # create_bar_plot(exact_counter, "en_exact")
    bar_plot(exact_counter, fixed_counter, decreasing_counter, "en_comparison")



    """versão francesa"""
    # limpa contadores, desnecessário estar a criar novos contadores
    exact_counter.clear()
    fixed_counter.clear()
    decreasing_counter.clear()

    # conta letras
    exact_count(exact_counter, "livros/3_MUSK_FR.txt")  # exact counter
    for x in range(quant):
        # limpra contadores
        fixed_counter.clear()
        decreasing_counter.clear()

        fixed_count(fixed_counter, "livros/3_MUSK_FR.txt")  # fixed counter with k value
        decreasing_count(decreasing_counter, "livros/3_MUSK_FR.txt")  # decreasing counter with increasing k value
        
        # verifica letras e ordem das letras são as mesmas da contagem exata
        verify_letters(fr_checker, exact_counter, fixed_counter, decreasing_counter)
    
    # mostra chances
    print("---------------------Versão fracesa---------------------")
    print("Para approximate counting with fixed chance:")
    print(f"    chance de as letras serem diferentes: {(fr_checker.fixed_order / quant) * 100}%")
    print(f"    chance de ordem das letras ser diferente: {(fr_checker.fixed_same_letters / quant) * 100}%")
    print("Para approximate counting with decreasing chance:")
    print(f"    chance de as letras serem diferentes: {(fr_checker.dec_order / quant) * 100}%")
    print(f"    chance de ordem das letras ser diferente: {(fr_checker.dec_same_letters / quant) * 100}%")
    
    # escreve informação no xlsx
    write_info(exact_counter, fixed_counter, decreasing_counter, "fr_version")

    # create_bar_plot(exact_counter, "fr_exact")
    bar_plot(exact_counter, fixed_counter, decreasing_counter, "fr_comparison")

    exit()
